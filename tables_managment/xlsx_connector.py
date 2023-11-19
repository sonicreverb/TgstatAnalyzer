import openpyxl
import os
import re

from datetime import datetime
from openpyxl.styles import Font

OUTPUT_FILENAME = 'TgstatAnalyzer_output.xlsx'


# принимает на вход API class Post, возвращает словарь с необходимыми хар-ками публикации
def get_item_data(item):
    # название канала
    post_id_name = item['link'].split('/')[1]
    if post_id_name == "c":
        post_id_name = item['link'].split('/')[2]

    # ссылка на сообщение
    post_url = "https://" + item['link']

    # содержимое сообщения
    post_text = item['text']

    # ссылка на фото
    if item['media'].get('media_type') == 'mediaPhoto':
        post_image_url = item['media']['file_url']
        if not post_image_url:
            post_image_url = "TgstatAPI return null img source."
    else:
        post_image_url = 'None'

    # ссылка на видео
    if item['media'].get('media_type', None) == 'mediaDocument' and item['media'].get('mime_type', '0') == 'video/mp4':
        post_video_url = item['media']['file_url']
        if not post_video_url:
            post_video_url = "TgstatAPI return null video source."
    else:
        post_video_url = 'None'

    # ссылка на источник
    post_source_link = post_url[:post_url.rfind('/')]

    # ссылки на сторонние ресурсы
    url_pattern = re.compile(r'<a\b[^>]*href=[\'"]?([^\'" >]+)[^>]*>(.*?)</a>|https?://\S+')
    matches = re.finditer(url_pattern, post_text)
    links = [match.group(1) if match.group(1) else match.group(0) for match in matches]

    post_rel_resourses = ''
    for link in links:
        if link[:-1] == "'":
            link = link[:-1]
        post_rel_resourses += link + '\n'

    if post_rel_resourses == '':
        post_rel_resourses = 'None'

    # дата публикации
    timestamp = item['date']
    dt_object = datetime.utcfromtimestamp(timestamp)
    post_date = dt_object.strftime('%d/%m/%Y %H:%M:%S')

    return {'postSourseName': post_id_name, 'postURL': post_url, 'postText': post_text, 'postImgURL': post_image_url,
            'postVideoURL': post_video_url, 'postSourceURL': post_source_link, 'relevResoursesURL': post_rel_resourses,
            'postPublishDate': post_date}


def get_column_data(col_num):
    if os.path.exists(OUTPUT_FILENAME):
        # если файл существует, открываем его и выбираем активный лист
        wb = openpyxl.load_workbook(OUTPUT_FILENAME)
        ws = wb.active
        # считывание уже присутствующих ссылок в таблице
        result = [ws.cell(row=row, column=col_num).value for row in range(1, ws.max_row + 1)]
    else:
        result = []

    return result


def write_to_excel(filename, data):
    # проверяем, существует ли файл
    if os.path.exists(filename):
        # если файл существует, открываем его и выбираем активный лист
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        # считывание уже присутствующих ссылок в таблице
        column_number = 3
        unique_text_in_table = [ws.cell(row=row, column=column_number).value for row in range(1, ws.max_row + 1)]
    else:
        # если файл не существует, создаем новую книгу Excel и выбираем активный лист
        wb = openpyxl.Workbook()
        ws = wb.active
        unique_text_in_table = []

    data_to_write = []
    for item in data:
        if item.get('postText') not in unique_text_in_table:
            data_to_write.append(item)
            unique_text_in_table.append(item.get('postText'))

    if not data_to_write:
        print(f'[ЗАПИСЬ В ЭКСЕЛЬ] Не найдено новых уникальных публикаций для запись в {filename}')
        return

    # добавляем заголовки таблицы
    headers = list(data_to_write[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    hyperlink_style = Font(color='0000FF', underline='single')

    # записываем данные в конец таблицы
    for row_num, row_data in enumerate(data_to_write, ws.max_row + 1):
        for col_num, key in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[key])
            if col_num in [2, 4, 5, 6, 7] and row_data[key] != 'None' and 'TgstatAPI return null' not in row_data[key]:
                cell.font = hyperlink_style
                cell.hyperlink = row_data[key]

    # сохраняем книгу Excel
    wb.save(filename)

    print(f'[ЗАПИСЬ В ЭКСЕЛЬ] Данные успешно записаны в файл {filename} [{len(data_to_write)}/{len(data)}]')
